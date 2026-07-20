# -*- coding: utf-8 -*-
"""测试用例 Markdown → Excel 转换工具

统一的 xlsx 生成方法，供 test-design 工作流调用。

用法：
    # 命令行
    python utils/testcase_to_xlsx.py <md_file> [--output <xlsx_file>]

    # Python 调用
    from utils.testcase_to_xlsx import convert_md_to_xlsx
    convert_md_to_xlsx("tests/data/test-cases-final.md", "tests/data/test-cases.xlsx")

输入格式要求：
    - Markdown 文件包含表格，表头必须包含：用例编号、用例标题、优先级、前置条件、测试步骤、预期结果
    - 支持"用例ID"自动映射为"用例编号"，"操作步骤"自动映射为"测试步骤"
    - 优先级列值：P0 / P1 / P2

输出格式：
    - Sheet 1「测试用例」：用例编号、用例标题、优先级、前置条件、测试步骤、预期结果、执行情况
    - Sheet 2「统计概览」：优先级分布、模块分布、执行情况统计
    - 优先级颜色：P0=粉色、P1=橙色、P2=绿色
    - 执行情况下拉：未执行/通过/失败/阻塞/跳过
"""

import re
import sys
from pathlib import Path
from typing import List, Dict, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation


# ── 样式常量 ──────────────────────────────────────────────────────────────────

HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

CELL_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="top")

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 优先级颜色
PRIORITY_COLORS = {
    "P0": PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid"),
    "P1": PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid"),
    "P2": PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"),
}

# 执行情况颜色
EXECUTION_COLORS = {
    "通过": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
    "失败": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    "阻塞": PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
    "跳过": PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"),
}

# 表头列定义
HEADERS = ['用例编号', '用例标题', '优先级', '前置条件', '测试步骤', '预期结果', '执行情况']

# 列宽定义
COLUMN_WIDTHS = {
    'A': 18,  # 用例编号
    'B': 40,  # 用例标题
    'C': 10,  # 优先级
    'D': 40,  # 前置条件
    'E': 50,  # 测试步骤
    'F': 50,  # 预期结果
    'G': 15,  # 执行情况
}

# 字段名映射（兼容不同格式的 md 表头）
FIELD_MAPPING = {
    '用例ID': '用例编号',
    '操作步骤': '测试步骤',
    '步骤': '测试步骤',
    '预期': '预期结果',
    '结果': '预期结果',
}


# ── 解析函数 ──────────────────────────────────────────────────────────────────

def parse_md_table(md_content: str) -> List[Dict[str, str]]:
    """解析 markdown 表格，返回用例列表

    支持：
    - 多个表格（按模块分组）
    - 字段名自动映射
    - 跳过分隔行

    :param md_content: markdown 文件内容
    :return: 用例字典列表
    """
    lines = md_content.strip().split('\n')
    cases = []
    i = 0

    while i < len(lines):
        line = lines[i].strip()

        # 找到表格开始位置（包含 | 的行，且是用例表头）
        if line.startswith('|') and any(key in line for key in ['用例ID', '用例编号']):
            # 解析表头
            headers = [h.strip() for h in line.split('|') if h.strip()]

            # 字段名标准化
            headers = [FIELD_MAPPING.get(h, h) for h in headers]

            # 跳过分隔行
            i += 2

            # 解析数据行
            while i < len(lines):
                data_line = lines[i].strip()
                if not data_line.startswith('|'):
                    break

                values = [v.strip() for v in data_line.split('|') if v.strip()]
                if len(values) >= len(headers):
                    case = {}
                    for j, header in enumerate(headers):
                        if j < len(values):
                            case[header] = values[j]
                    # 确保必要字段存在
                    if case.get('用例编号'):
                        cases.append(case)
                i += 1
        else:
            i += 1

    return cases


# ── 生成函数 ──────────────────────────────────────────────────────────────────

def create_xlsx(cases: List[Dict[str, str]], output_path: Path) -> None:
    """创建 xlsx 文件

    :param cases: 用例字典列表
    :param output_path: 输出文件路径
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"

    # 执行情况下拉选项
    execution_validation = DataValidation(
        type="list",
        formula1='"未执行,通过,失败,阻塞,跳过"',
        allow_blank=True
    )
    execution_validation.error = "请选择有效的执行情况"
    execution_validation.errorTitle = "无效输入"
    ws.add_data_validation(execution_validation)

    # 写入表头
    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # 写入数据
    for row_idx, case in enumerate(cases, 2):
        # 用例编号
        ws.cell(row=row_idx, column=1, value=case.get('用例编号', '')).border = THIN_BORDER

        # 用例标题
        ws.cell(row=row_idx, column=2, value=case.get('用例标题', '')).border = THIN_BORDER

        # 优先级（带颜色）
        priority = case.get('优先级', '')
        priority_cell = ws.cell(row=row_idx, column=3, value=priority)
        priority_cell.border = THIN_BORDER
        priority_cell.alignment = CENTER_ALIGNMENT
        if priority in PRIORITY_COLORS:
            priority_cell.fill = PRIORITY_COLORS[priority]

        # 前置条件
        ws.cell(row=row_idx, column=4, value=case.get('前置条件', '')).border = THIN_BORDER

        # 测试步骤
        ws.cell(row=row_idx, column=5, value=case.get('测试步骤', '')).border = THIN_BORDER

        # 预期结果
        ws.cell(row=row_idx, column=6, value=case.get('预期结果', '')).border = THIN_BORDER

        # 执行情况（默认未执行）
        execution_cell = ws.cell(row=row_idx, column=7, value='未执行')
        execution_cell.border = THIN_BORDER
        execution_cell.alignment = CENTER_ALIGNMENT
        execution_validation.add(execution_cell)

        # 设置对齐方式
        for col in range(1, 7):
            ws.cell(row=row_idx, column=col).alignment = CELL_ALIGNMENT

    # 设置列宽
    for col_letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # 冻结首行
    ws.freeze_panes = 'A2'

    # 创建统计 sheet
    _create_stats_sheet(wb, cases)

    # 保存文件
    wb.save(output_path)


def _create_stats_sheet(wb: Workbook, cases: List[Dict[str, str]]) -> None:
    """创建统计概览 sheet"""
    ws_stats = wb.create_sheet("统计概览")

    # 统计优先级分布
    priority_count = {'P0': 0, 'P1': 0, 'P2': 0}
    for case in cases:
        p = case.get('优先级', '')
        if p in priority_count:
            priority_count[p] += 1

    # 写入优先级统计
    ws_stats.cell(row=1, column=1, value="优先级分布").font = Font(bold=True, size=12)
    ws_stats.cell(row=2, column=1, value="优先级")
    ws_stats.cell(row=2, column=2, value="数量")
    ws_stats.cell(row=2, column=3, value="占比")

    total = len(cases)
    for row, (priority, count) in enumerate(priority_count.items(), 3):
        ws_stats.cell(row=row, column=1, value=priority)
        ws_stats.cell(row=row, column=2, value=count)
        ws_stats.cell(row=row, column=3, value=f"{count/total*100:.1f}%" if total > 0 else "0%")

    ws_stats.cell(row=6, column=1, value="总计").font = Font(bold=True)
    ws_stats.cell(row=6, column=2, value=total).font = Font(bold=True)
    ws_stats.cell(row=6, column=3, value="100%").font = Font(bold=True)

    # 统计模块分布
    module_count = {}
    for case in cases:
        case_id = case.get('用例编号', '')
        # 从编号提取模块名（如 TC-BUBBLE-001 → BUBBLE）
        parts = case_id.split('-')
        module = parts[1] if len(parts) >= 3 else '其他'
        module_count[module] = module_count.get(module, 0) + 1

    ws_stats.cell(row=8, column=1, value="模块分布").font = Font(bold=True, size=12)
    ws_stats.cell(row=9, column=1, value="模块")
    ws_stats.cell(row=9, column=2, value="数量")

    for row, (module, count) in enumerate(sorted(module_count.items()), 10):
        ws_stats.cell(row=row, column=1, value=module)
        ws_stats.cell(row=row, column=2, value=count)

    # 执行情况统计
    exec_row = 10 + len(module_count) + 2
    ws_stats.cell(row=exec_row, column=1, value="执行情况统计").font = Font(bold=True, size=12)
    ws_stats.cell(row=exec_row + 1, column=1, value="执行情况")
    ws_stats.cell(row=exec_row + 1, column=2, value="数量")
    ws_stats.cell(row=exec_row + 1, column=3, value="占比")

    execution_stats = {'未执行': total, '通过': 0, '失败': 0, '阻塞': 0, '跳过': 0}
    for row, (status, count) in enumerate(execution_stats.items(), exec_row + 2):
        ws_stats.cell(row=row, column=1, value=status)
        ws_stats.cell(row=row, column=2, value=count)
        ws_stats.cell(row=row, column=3, value=f"{count/total*100:.1f}%" if total > 0 else "0%")

    # 设置列宽
    ws_stats.column_dimensions['A'].width = 15
    ws_stats.column_dimensions['B'].width = 10
    ws_stats.column_dimensions['C'].width = 10


# ── 主入口 ────────────────────────────────────────────────────────────────────

def convert_md_to_xlsx(md_path: str, xlsx_path: Optional[str] = None) -> Path:
    """将 markdown 测试用例转换为 xlsx

    :param md_path: markdown 文件路径
    :param xlsx_path: xlsx 输出路径（默认同目录同名）
    :return: xlsx 文件路径
    """
    md_path = Path(md_path)
    if not md_path.exists():
        raise FileNotFoundError(f"Markdown 文件不存在: {md_path}")

    if xlsx_path is None:
        xlsx_path = md_path.with_suffix('.xlsx')
    else:
        xlsx_path = Path(xlsx_path)

    # 读取 markdown
    with open(md_path, 'r', encoding='utf-8') as f:
        md_content = f.read()

    # 解析表格
    cases = parse_md_table(md_content)
    if not cases:
        raise ValueError(f"未找到测试用例数据: {md_path}")

    # 创建 xlsx
    create_xlsx(cases, xlsx_path)

    print(f"✅ xlsx 文件已生成: {xlsx_path}")
    print(f"   总用例数: {len(cases)}")

    # 统计优先级
    priority_count = {'P0': 0, 'P1': 0, 'P2': 0}
    for case in cases:
        p = case.get('优先级', '')
        if p in priority_count:
            priority_count[p] += 1
    print(f"   P0: {priority_count['P0']} | P1: {priority_count['P1']} | P2: {priority_count['P2']}")

    return xlsx_path


def main():
    """命令行入口"""
    import argparse

    parser = argparse.ArgumentParser(description="测试用例 Markdown → Excel 转换")
    parser.add_argument("md_file", help="Markdown 文件路径")
    parser.add_argument("--output", "-o", help="xlsx 输出路径（默认同目录同名）")
    args = parser.parse_args()

    try:
        convert_md_to_xlsx(args.md_file, args.output)
    except Exception as e:
        print(f"❌ 转换失败: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
